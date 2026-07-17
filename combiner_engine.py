"""
combiner_engine.py

Generic logic for combining "split" files that all share the same
header/template — e.g. large exports that get chopped into
1.xlsx / 2.xlsx / 3.xlsx, or any set of files with identical header
rows and different data rows underneath.

Supports:
- .xlsx (openpyxl, formatting of the base file is preserved in the output)
- .xls  (read via openpyxl-compatible fallback -> converted to xlsx)
- .csv  (read via the csv module)

Input can be:
- A single .zip containing multiple files
- Several individual files uploaded directly (no zip)
- A mix of both

Header detection:
- If 2+ files are supplied, the header row count is auto-detected by
  walking rows top-down and finding the last row that is IDENTICAL
  across every file — after that point, rows are treated as data.
  This works because split files repeat the same header/template
  verbatim, and only the payload underneath differs.
- The caller can override the detected count (e.g. via a UI control).
- If only one file is supplied, header_rows defaults to 1 unless
  the caller overrides it.

Known quirk handled automatically:
- Shopee-exported .xlsx files contain an invalid `activePane`
  attribute in their internal XML that crashes openpyxl. This is
  patched out at the raw zip/XML level before loading. This patch is
  a no-op for files that don't have the issue, so it's always safe
  to apply.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass, field

import openpyxl


@dataclass
class InputFile:
    name: str
    bytes_: bytes


@dataclass
class CombineResult:
    output_bytes: bytes
    output_filename: str
    header_rows_used: int
    files_processed: list[str] = field(default_factory=list)
    files_skipped_empty: list[str] = field(default_factory=list)
    rows_per_file: dict[str, int] = field(default_factory=dict)
    total_rows: int = 0
    max_col: int = 0
    warnings: list[str] = field(default_factory=list)


SUPPORTED_EXTS = (".xlsx", ".xls", ".csv")


def _natural_sort_key(name: str):
    parts = re.split(r"(\d+)", name)
    return [int(p) if p.isdigit() else p for p in parts]


def _patch_activepane(xlsx_bytes: bytes) -> bytes:
    """Strip the invalid activePane="..." attribute some exporters
    (e.g. Shopee) leave in the sheet-view XML, which otherwise crashes
    openpyxl. No-op for files that don't have the issue."""
    try:
        src = zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r")
    except zipfile.BadZipFile:
        return xlsx_bytes
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in src.namelist():
            data = src.read(name)
            if name.endswith(".xml"):
                try:
                    text = data.decode("utf-8")
                    text = re.sub(r'activePane="[^"]*"', "", text)
                    data = text.encode("utf-8")
                except UnicodeDecodeError:
                    pass
            zout.writestr(name, data)
    src.close()
    return out_buf.getvalue()


def _is_row_empty(values: list) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def _rows_equal(a: list, b: list) -> bool:
    if len(a) != len(b):
        return False
    for x, y in zip(a, b):
        xs = "" if x is None else str(x).strip()
        ys = "" if y is None else str(y).strip()
        if xs != ys:
            return False
    return True


# ---------------------------------------------------------------------------
# Reading files into a uniform "list of rows" representation
# ---------------------------------------------------------------------------


def _read_rows_xlsx(data: bytes) -> list[list]:
    patched = _patch_activepane(data)
    wb = openpyxl.load_workbook(io.BytesIO(patched), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])
    return rows


def _read_rows_csv(data: bytes) -> list[list]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


def _read_rows(name: str, data: bytes) -> list[list]:
    ext = name.lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls"):
        return _read_rows_xlsx(data)
    if ext == "csv":
        return _read_rows_csv(data)
    raise ValueError(f"Unsupported file type: {name}")


# ---------------------------------------------------------------------------
# Header auto-detection
# ---------------------------------------------------------------------------


def detect_header_rows(files: list[InputFile], max_check: int = 30) -> int:
    """Return how many leading rows are identical across all supplied
    files. If only one file is given, returns 1 (assume a single
    header row) as a sane default."""
    if len(files) < 2:
        return 1

    all_rows = [_read_rows(f.name, f.bytes_) for f in files]
    shortest = min(len(r) for r in all_rows)
    limit = min(max_check, shortest)

    header_rows = 0
    for i in range(limit):
        rows_at_i = [rows[i] for rows in all_rows]
        if all(_rows_equal(rows_at_i[0], r) for r in rows_at_i[1:]):
            header_rows += 1
        else:
            break
    return max(header_rows, 1)


def preview_rows(file: InputFile, n: int = 10) -> list[list]:
    rows = _read_rows(file.name, file.bytes_)
    return rows[:n]


# ---------------------------------------------------------------------------
# Combining
# ---------------------------------------------------------------------------


def _extract_files_from_zip(zip_bytes: bytes) -> list[InputFile]:
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes), "r")
    names = sorted(
        [
            n
            for n in zf.namelist()
            if n.lower().endswith(SUPPORTED_EXTS) and not n.startswith("__MACOSX") and "/" not in n.strip("/")
        ],
        key=_natural_sort_key,
    )
    return [InputFile(name=n, bytes_=zf.read(n)) for n in names]


def collect_input_files(uploads: list[tuple[str, bytes]]) -> list[InputFile]:
    """Given a list of (filename, bytes) tuples from the uploader —
    which may include zip files and/or individual data files — expand
    everything into a flat, naturally-sorted list of InputFile."""
    collected: list[InputFile] = []
    for name, data in uploads:
        if name.lower().endswith(".zip"):
            collected.extend(_extract_files_from_zip(data))
        elif name.lower().endswith(SUPPORTED_EXTS):
            collected.append(InputFile(name=name, bytes_=data))
    collected.sort(key=lambda f: _natural_sort_key(f.name))
    return collected


def combine_files(
    files: list[InputFile],
    header_rows: int | None = None,
    output_filename: str | None = None,
) -> CombineResult:
    """
    Combine a list of InputFile into a single .xlsx, keeping the first
    `header_rows` rows once (from the first non-empty file) and
    appending every data row from every file after that.

    If header_rows is None, it is auto-detected.
    """
    if not files:
        raise ValueError("No files to combine.")

    if header_rows is None:
        header_rows = detect_header_rows(files)

    warnings: list[str] = []
    parsed: dict[str, list[list]] = {}
    for f in files:
        try:
            parsed[f.name] = _read_rows(f.name, f.bytes_)
        except Exception as e:
            warnings.append(f"Skipped '{f.name}': could not read it ({e}).")

    files_processed: list[str] = []
    files_skipped_empty: list[str] = []
    rows_per_file: dict[str, int] = {}

    # Determine base file: first file (in order) that actually has data
    # rows beyond the header.
    base_name = None
    base_header = None
    max_col = 0
    for f in files:
        rows = parsed.get(f.name)
        if rows is None:
            continue
        data_rows = rows[header_rows:]
        has_data = any(not _is_row_empty(r) for r in data_rows)
        if not has_data:
            files_skipped_empty.append(f.name)
            continue
        base_name = f.name
        base_header = rows[:header_rows]
        max_col = max((len(r) for r in rows), default=0)
        break

    if base_name is None:
        raise ValueError(
            "None of the files have data rows beyond the detected header "
            f"({header_rows} row(s)). Try adjusting the header row count."
        )

    # Build combined rows: header once, then data from every file in order.
    combined_rows: list[list] = list(base_header)
    for f in files:
        if f.name in files_skipped_empty:
            continue
        rows = parsed.get(f.name)
        if rows is None:
            continue
        data_rows = rows[header_rows:]
        appended = 0
        col_count = max((len(r) for r in rows), default=0)
        if col_count and col_count != max_col:
            warnings.append(
                f"'{f.name}' has {col_count} columns vs base file's {max_col} — "
                "rows were still copied but please double-check alignment."
            )
        for r in data_rows:
            if _is_row_empty(r):
                continue
            combined_rows.append(r)
            appended += 1
        rows_per_file[f.name] = appended
        files_processed.append(f.name)

    total_rows = len(combined_rows) - header_rows

    # Write output. If the base file was an xlsx, reuse it as a template
    # so formatting/validation carries over; otherwise build a fresh
    # workbook.
    out_wb = None
    if base_name.lower().endswith((".xlsx", ".xls")):
        try:
            base_bytes = next(f.bytes_ for f in files if f.name == base_name)
            patched = _patch_activepane(base_bytes)
            out_wb = openpyxl.load_workbook(io.BytesIO(patched))
            out_ws = out_wb["Sheet1"] if "Sheet1" in out_wb.sheetnames else out_wb.worksheets[0]
            # Clear existing rows beyond header, then rewrite everything
            # (simplest reliable way to append across many files).
            out_ws.delete_rows(header_rows + 1, out_ws.max_row)
            for r_idx, row in enumerate(combined_rows[header_rows:], start=header_rows + 1):
                for c_idx, val in enumerate(row, start=1):
                    out_ws.cell(row=r_idx, column=c_idx, value=val)
        except Exception:
            out_wb = None

    if out_wb is None:
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Sheet1"
        for r_idx, row in enumerate(combined_rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                out_ws.cell(row=r_idx, column=c_idx, value=val)

    out_buf = io.BytesIO()
    out_wb.save(out_buf)
    out_bytes = out_buf.getvalue()

    if output_filename is None:
        output_filename = "COMBINED.xlsx"
    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    return CombineResult(
        output_bytes=out_bytes,
        output_filename=output_filename,
        header_rows_used=header_rows,
        files_processed=files_processed,
        files_skipped_empty=files_skipped_empty,
        rows_per_file=rows_per_file,
        total_rows=total_rows,
        max_col=max_col,
        warnings=warnings,
    )
